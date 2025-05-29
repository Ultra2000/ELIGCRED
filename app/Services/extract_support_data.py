import pandas as pd
from openpyxl import load_workbook
from pathlib import Path
import datetime

def excel_date_to_str(excel_date):
    if not excel_date:
        return ''
    if isinstance(excel_date, datetime.datetime):
        return excel_date.strftime('%Y-%m-%d')
    try:
        return (datetime.datetime(1899, 12, 30) + datetime.timedelta(days=float(excel_date))).strftime('%Y-%m-%d')
    except Exception:
        return str(excel_date)

# Chemin du fichier source et du fichier de sortie
SUPPORT_FILE = Path('SUPPORT_CREDIT_IND.xlsx')
OUTPUT_FILE = Path('support_data_for_ia.xlsx')

wb = load_workbook(SUPPORT_FILE, data_only=True)
ws = wb.active

data = {
    'numero_dossier': ws['D6'].value or '',
    'guichet': ws['D7'].value or '',
    'code_adherent': ws['D8'].value or '',
    'date_prise_info': excel_date_to_str(ws['D9'].value),
    'renouvellement': ws['D10'].value or '',
    'activite_financer': ws['D11'].value or '',
    'montant_sollicite': ws['A13'].value or 0,
    'duree_sollicitee': ws['C13'].value or 0,
    'periodicite_sollicitee': ws['E13'].value or '',
    'nom_client': ws['B17'].value or '',
    'prenom_client': ws['B18'].value or '',
    'sexe': ws['D19'].value or '',
    'date_naissance': excel_date_to_str(ws['B20'].value),
    'lieu_naissance': ws['F20'].value or '',
    'adresse': ws['B21'].value or '',
    'residence_depuis': ws['F21'].value or '',
    'metier': ws['B22'].value or '',
    'activites': ws['F22'].value or '',
    'numero_ifu': ws['C24'].value or '',
    'situation_matrimoniale': ws['D26'].value or '',
    'personnes_charge': ws['D27'].value or 0,
    'reference': ws['D30'].value or '',
    'adresse_entreprise': ws['D33'].value or '',
    'niveau_concurrence': ws['D40'].value or '',
    'salaire_net': ws['C45'].value or 0,
    'total_revenus': ws['C53'].value or 0,
    'total_depenses': ws['F53'].value or 0,
    'actif_net': ws['F54'].value or 0,
    'total_credits': ws['D67'].value or 0,
    'cout_total_projet': ws['F80'].value or 0,
    'besoin_financement': ws['C85'].value or 0,
    'total_garanties': ws['C96'].value or 0,
    'encaisse': ws['C101'].value or 0,
    'total_actif': ws['C112'].value or 0,
    'total_passif': ws['F112'].value or 0,
    'vente': ws['C118'].value or 0,
    'surplus_net': ws['C130'].value or 0,
    'statut': 'FAVORABLE',  # Valeur par défaut pour le statut
    'age': 0,  # Valeur par défaut pour l'âge
    'score_ia': 0,  # Valeur par défaut pour le score IA
    'montant_propose': 0,  # Valeur par défaut pour le montant proposé
    'duree_proposee': 0  # Valeur par défaut pour la durée proposée
}

# Nettoyage et conversion
for k in ['montant_sollicite','duree_sollicitee','salaire_net','total_revenus','total_depenses','actif_net','total_credits','cout_total_projet','besoin_financement','total_garanties','encaisse','total_actif','total_passif','vente','surplus_net','personnes_charge']:
    try:
        data[k] = float(str(data[k]).replace(' ', '').replace(',', '.')) if data[k] is not None else 0
    except Exception:
        data[k] = 0

# Créer un DataFrame et sauvegarder
pd.DataFrame([data]).to_excel(OUTPUT_FILE, index=False)
print(f"Extraction terminée. Données sauvegardées dans {OUTPUT_FILE}") 